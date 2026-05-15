"""
BizFit 찾검 키워드 업데이트 자동화 (엑셀 입출력)
=============================================
사용법:
  1. 키워드_변경.xlsx 파일에 MID, 캠페인명, 키워드, 메모 입력
  2. python BizFit_키워드_변경_MID.py 실행
  3. 같은 파일에 처리결과/처리시각 열이 자동으로 채워짐

키워드_변경.xlsx 구조:
  A열: MID (상품MID/PID)
  B열: 캠페인명 (선택)
  C열: 키워드 (쉼표로 구분, 예: 귀금속,귀금속방,귀금속위치)
  D열: 메모 (선택, 비워두면 오늘 날짜(MM-DD) 자동 입력)
  E열: 처리결과 → 자동 입력
  F열: 처리시각 → 자동 입력
"""

import os
import time
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import openpyxl
from openpyxl.styles import PatternFill, Font

# ============================================================
#  설정 영역 - 여기만 수정하세요
# ============================================================
CONFIG = {
    "url":          "https://admin.bizfit.kr/v1/admin/pages/login/controller/",
    "id":           "higherad",
    "password":     "hi1107",
    "headless":     False,
    "excel_file":   os.path.join(os.path.dirname(os.path.abspath(__file__)), "키워드_변경_MID.xlsx"),
}
# ============================================================

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FONT = Font(color="276221", bold=True)
RED_FONT   = Font(color="9C0006", bold=True)


def log(msg):
    now = datetime.now().strftime("%H:%M:%S")
    print(f"[{now}] {msg}")


def create_sample_excel(filepath):
    """샘플 엑셀 파일 생성"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "키워드업데이트목록"

    headers = ["MID", "캠페인명", "키워드", "메모", "처리결과", "처리시각"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    ws.cell(row=2, column=1, value="1839711962")
    ws.cell(row=2, column=2, value="예시_캠페인명")
    ws.cell(row=2, column=3, value="귀금속,귀금속방,귀금속위치")
    ws.cell(row=2, column=4, value="")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    wb.save(filepath)
    log(f"'{filepath}' 샘플 파일을 생성했어요.")
    log("MID, 캠페인명, 키워드를 입력한 후 다시 실행해주세요.")


def load_entries_from_excel(filepath):
    """엑셀에서 MID, 캠페인명, 키워드 읽기 (1행=헤더 스킵)"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        mid_cell     = row[0]
        name_cell    = row[1] if len(row) > 1 else None
        keyword_cell = row[2] if len(row) > 2 else None
        memo_cell    = row[3] if len(row) > 3 else None

        mid      = str(mid_cell.value).strip() if mid_cell.value else ""
        name     = str(name_cell.value).strip() if (name_cell and name_cell.value) else ""
        keywords = str(keyword_cell.value).strip() if (keyword_cell and keyword_cell.value) else ""
        memo     = str(memo_cell.value).strip() if (memo_cell and memo_cell.value) else ""

        if mid:
            entries.append({
                "mid":      mid,
                "name":     name,
                "keywords": keywords,
                "memo":     memo,
                "row":      mid_cell.row,
            })

    return wb, ws, entries


def write_result(ws, row_num, success, timestamp):
    """엑셀 결과 열에 성공/실패 기록 (E열=결과, F열=시각)"""
    result_cell = ws.cell(row=row_num, column=5)
    time_cell   = ws.cell(row=row_num, column=6)

    if success:
        result_cell.value = "✅ 성공"
        result_cell.fill  = GREEN_FILL
        result_cell.font  = GREEN_FONT
    else:
        result_cell.value = "❌ 실패"
        result_cell.fill  = RED_FILL
        result_cell.font  = RED_FONT

    time_cell.value = timestamp


def navigate_to_campaign_list(page):
    """캠페인 관리 페이지로 직접 이동 후 진행중 필터 클릭"""
    CAMPAIGN_URL = "https://admin.bizfit.kr/v1/admin/pages/adman/controller/ad_manage2.php"
    log("  캠페인 관리 페이지 직접 이동...")
    page.goto(CAMPAIGN_URL, wait_until="networkidle")
    time.sleep(1.5)

    log("  진행중 버튼 클릭...")
    page.locator("button.status[data-val='3']").click()
    page.wait_for_load_state("networkidle")
    time.sleep(1.5)


def search_by_mid(page, mid):
    """검색항목을 상품MID(PID)로 변경 후 MID 검색"""
    log("  검색항목 → 상품MID(PID) 변경...")
    page.locator("select#category").select_option(value="a.mid")
    time.sleep(0.4)

    log(f"  MID [{mid}] 검색...")
    page.locator("input.form-control[name='keyword']").fill(mid)
    time.sleep(0.3)

    page.get_by_role("button", name="조회하기").click()
    page.wait_for_load_state("networkidle")
    time.sleep(2)


def get_row_pcodes(page):
    """리스트에서 pcode 목록 추출 (보기 버튼의 onclick에서 파싱)"""
    pcodes = []
    buttons = page.locator("button.btn.btn-outline-ocr").all()
    for btn in buttons:
        onclick = btn.get_attribute("onclick") or ""
        if "modal_ad_edit.php?pcode=" in onclick:
            pcode = onclick.split("pcode=")[1].split("'")[0].split('"')[0]
            if pcode not in pcodes:
                pcodes.append(pcode)
    return pcodes


def update_keyword_popup(page, pcode, keyword_type, keywords_raw, label):
    """
    찾검 키워드 팝업 열기 → 기존 키워드 모두 삭제 → 첫 번째 input에 콤마 포함 통째로 입력 → 저장
    keyword_type: 'A' or 'B'
    keywords_raw: '귀금속,귀금속방,귀금속위치' (콤마 그대로)
    """
    popup_url = (
        f"https://admin.bizfit.kr/v1/admin/pages/adman/controller/"
        f"modal_keyword_view_np.php?pcode={pcode}&sch_keyword_type={keyword_type}"
    )

    log(f"  [{label}] 찾검 키워드{keyword_type} 팝업 열기...")

    keyword_page = page.context.new_page()
    keyword_page.goto(popup_url, wait_until="networkidle")
    time.sleep(1.5)

    try:
        inputs = keyword_page.locator("input[name='base_keyword[]']").all()
        log(f"  [{label}] 기존 키워드 행 수: {len(inputs)}")

        for inp in inputs:
            inp.fill("")

        time.sleep(0.3)

        if inputs:
            inputs[0].fill(keywords_raw)
        else:
            log(f"  [{label}] ⚠️ 입력 행을 찾을 수 없음")

        time.sleep(0.3)

        keyword_page.on("dialog", lambda dialog: dialog.accept())
        keyword_page.locator("button#btn_submit").click()
        keyword_page.wait_for_load_state("networkidle")
        time.sleep(1.0)

        log(f"  [{label}] ✅ 찾검 키워드{keyword_type} 저장 완료")
        keyword_page.close()
        return True

    except PlaywrightTimeout:
        log(f"  [{label}] ❌ 키워드{keyword_type} 시간 초과")
        keyword_page.close()
        return False
    except Exception as e:
        log(f"  [{label}] ❌ 키워드{keyword_type} 오류: {e}")
        keyword_page.close()
        return False


def write_memo(page, pcode, label, memo_text=""):
    """메모 페이지에 memo_text 등록 (비어있으면 오늘 날짜 MM-DD 사용)"""
    memo_url = (
        f"https://admin.bizfit.kr/v1/admin/pages/adman/controller/"
        f"ad_memo.php?pcode={pcode}"
    )
    text = memo_text if memo_text else datetime.now().strftime("%m-%d")

    log(f"  [{label}] 메모 등록 중 ({text})...")

    memo_page = page.context.new_page()
    memo_page.goto(memo_url, wait_until="networkidle")
    time.sleep(1.0)

    try:
        memo_page.wait_for_selector("textarea#memo", state="visible", timeout=8_000)
        memo_page.locator("textarea#memo").fill(text)
        time.sleep(0.3)

        memo_page.on("dialog", lambda dialog: dialog.accept())
        memo_page.locator("button.btn-outline-danger").click()
        memo_page.wait_for_load_state("networkidle")
        time.sleep(1.0)

        log(f"  [{label}] ✅ 메모 등록 완료")
        memo_page.close()
        return True

    except PlaywrightTimeout:
        log(f"  [{label}] ❌ 메모 페이지 시간 초과")
        memo_page.close()
        return False
    except Exception as e:
        log(f"  [{label}] ❌ 메모 등록 오류: {e}")
        memo_page.close()
        return False


def update_campaign_keywords(page, mid, name, keywords_raw, entry_memo=""):
    """단일 MID 키워드 업데이트 전체 흐름"""
    label = f"{mid}" + (f" ({name})" if name else "")

    if not keywords_raw:
        log(f"  [{label}] ⚠️ 키워드가 비어있어요. 건너뜀.")
        return False

    log(f"  [{label}] 키워드: {keywords_raw}")

    # 캠페인 관리 페이지로 이동 후 MID 검색
    navigate_to_campaign_list(page)
    search_by_mid(page, mid)

    # 리스트에서 pcode 수집
    pcodes = get_row_pcodes(page)

    if not pcodes:
        log(f"  [{label}] ⚠️ 검색 결과 없음 (MID: {mid})")
        return False

    log(f"  [{label}] {len(pcodes)}개 캠페인 발견: {pcodes}")

    all_ok = True

    for pcode in pcodes:
        pcode_label = f"{label} / {pcode}"

        # 찾검 키워드A 업데이트
        ok_a = update_keyword_popup(page, pcode, "A", keywords_raw, pcode_label)
        time.sleep(0.5)

        # 찾검 키워드B 업데이트
        ok_b = update_keyword_popup(page, pcode, "B", keywords_raw, pcode_label)
        time.sleep(0.5)

        if not (ok_a and ok_b):
            log(f"  [{pcode_label}] ❌ 키워드 업데이트 일부 실패 (A:{ok_a}, B:{ok_b})")
            all_ok = False
            continue

        # 수정하기 버튼 클릭
        log(f"  [{pcode_label}] 수정하기 버튼 클릭...")
        main_url = f"https://admin.bizfit.kr/v1/admin/pages/adman/controller/modal_ad_edit.php?pcode={pcode}"
        page.goto(main_url, wait_until="networkidle")
        time.sleep(1.0)

        try:
            page.locator("button#mod_btn").click()
            page.wait_for_load_state("networkidle")
            time.sleep(1.5)
            log(f"  [{pcode_label}] ✅ 수정하기 완료")
        except Exception as e:
            log(f"  [{pcode_label}] ❌ 수정하기 오류: {e}")
            all_ok = False
            continue

        # 메모 등록
        memo_ok = write_memo(page, pcode, pcode_label, entry_memo)
        if not memo_ok:
            all_ok = False

    return all_ok


def run():
    filepath = CONFIG["excel_file"]

    if not os.path.exists(filepath):
        create_sample_excel(filepath)
        time.sleep(5)
        return

    wb, ws, entries = load_entries_from_excel(filepath)
    if not entries:
        log("처리할 MID가 없어요. 엑셀 파일에 MID를 입력해주세요.")
        return

    log(f"총 {len(entries)}개 캠페인 키워드 업데이트 시작")
    log("=" * 40)

    success_count = 0
    fail_count    = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=CONFIG["headless"])
        context = browser.new_context()
        page    = context.new_page()

        # ── 로그인 ─────────────────────────────────────────
        log("로그인 중...")
        page.goto(CONFIG["url"], wait_until="networkidle")
        page.wait_for_selector("#aid", state="visible", timeout=10_000)
        time.sleep(0.5)
        page.locator("#aid").fill(CONFIG["id"])
        page.locator("input[type='password']").fill(CONFIG["password"])
        time.sleep(0.5)
        page.locator("input[type='button'][value='LOGIN']").click()
        page.wait_for_load_state("networkidle")
        time.sleep(1.5)
        log("로그인 완료")
        log("=" * 40)

        # ── 각 MID 처리 ────────────────────────────────────
        for i, entry in enumerate(entries, 1):
            log(f"[{i}/{len(entries)}] MID: {entry['mid']} 처리 중...")
            try:
                success = update_campaign_keywords(
                    page,
                    entry["mid"],
                    entry["name"],
                    entry["keywords"],
                    entry.get("memo", ""),
                )
            except Exception as e:
                log(f"  ❌ 예외 발생: {e}")
                success = False

            timestamp = datetime.now().strftime("%H:%M:%S")

            write_result(ws, entry["row"], success, timestamp)
            wb.save(filepath)

            if success:
                success_count += 1
            else:
                fail_count += 1

            time.sleep(1)

        browser.close()

    # ── 결과 요약 ──────────────────────────────────────────
    log("=" * 40)
    log(f"✅ 성공: {success_count}개")
    log(f"❌ 실패: {fail_count}개")
    log(f"📄 결과가 '{filepath}' 에 저장되었어요.")
    log("=" * 40)
    log("모든 작업 완료!")


if __name__ == "__main__":
    try:
        run()
    except PlaywrightTimeout as e:
        log(f"시간 초과: {e}")
        time.sleep(10)
    except Exception as e:
        log(f"오류 발생: {e}")
        time.sleep(10)
        raise
